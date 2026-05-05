import csv
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


REPO = Path(r"D:\DOCUMENTS\GitHub\HUNTS")
XLSX_DIR = REPO / "pipeline" / "raw" / "2026" / "xlsx"
OUT = REPO / "pipeline" / "normalized" / "2026_hunt_tables_normalized.csv"

# Files user confirmed as source of truth
SOURCE_FILES = [
    "2026 Elk.xlsx",
    "2026_ANTLERLESS_ELK.xlsx",
    "2026_BUCK_PRONGHORN.xlsx",
    "DOE_PRONGHORN.xlsx",
    "BUCK DEER.xlsx",
    "ANTLERLESS_MOOSE.xlsx",
    "deer hunter's choice.xlsx",
    "2026_MOUNTAIN_GOAT.xlsx",
    "2026_DESERT_BIGHORN.xlsx",
    "2026_ROCKY_MOUNTAIN_SHEEP.xlsx",
    "2026_FALL_TURKEY_EITHER_SEX.xlsx",
    "2026_BEARDED_TURKEY.xlsx",
    "2026_BLACK_BEAR.xlsx",
    "COUGAR STATEWIDE.xlsx",
]

CANON = [
    "hunt_name",
    "hunt_code",
    "sex_type",
    "species",
    "weapon",
    "hunt_type",
    "season",
    "res_permits",
    "non_res_permits",
    "total_permits",
    "permits_2025_res",
    "permits_2025_nr",
    "permits_2025_total",
    "source_file",
    "source_sheet",
]


def norm_header(s: str) -> str:
    return s.strip().lower().replace("_", "-")


def infer_header_row(first_rows) -> int:
    best_row = 1
    best_score = -1
    for r, vals in enumerate(first_rows, start=1):
        txt = [str(v).strip().lower() for v in vals if v is not None and str(v).strip()]
        score = 0
        for t in txt:
            if t in {"hunt_name", "hunt-name", "hunt number", "hunt-number", "species", "weapon", "season"}:
                score += 2
            if "permit" in t:
                score += 1
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def parse_sheet(path: Path, ws):
    # Special-case: single-row no-header files (e.g., COUGAR STATEWIDE.xlsx)
    if path.name.lower() == "cougar statewide.xlsx":
        vals = [ws.cell(row=1, column=c).value for c in range(1, 20)]
        txt_nonempty = [v for v in vals if v is not None and str(v).strip()]
        if txt_nonempty:
            row = {
                "hunt_name": "" if vals[0] is None else str(vals[0]).strip(),
                "hunt_code": "" if vals[1] is None else str(vals[1]).strip(),
                "sex_type": "" if vals[2] is None else str(vals[2]).strip(),
                "species": "" if vals[3] is None else str(vals[3]).strip(),
                "weapon": "" if vals[4] is None else str(vals[4]).strip(),
                "hunt_type": "" if vals[5] is None else str(vals[5]).strip(),
                "season": "" if vals[6] is None else str(vals[6]).strip(),
                "permits_2025_res": "",
                "permits_2025_nr": "",
                "permits_2025_total": "",
                "source_file": path.name,
                "source_sheet": ws.title,
            }
            row["res_permits"] = row["permits_2025_res"]
            row["non_res_permits"] = row["permits_2025_nr"]
            row["total_permits"] = row["permits_2025_total"]
            return [row]
        return []

    first_rows = []
    for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=40, values_only=True):
        first_rows.append(list(row))
    hr = infer_header_row(first_rows)

    header_cells = first_rows[hr - 1] if hr - 1 < len(first_rows) else []
    headers = [norm_header(str(v)) if v is not None else "" for v in header_cells]
    idx = {h: i for i, h in enumerate(headers) if h}

    # BUCK DEER.xlsx sometimes has no explicit header row.
    # Fallback map by position from known schema.
    fallback_positions = {
        "hunt_name": 0,
        "hunt-number": 1,
        "sex-type": 2,
        "species": 3,
        "weapon": 4,
        "hunt-type": 5,
        "season": 6,
        "2025_res_permit": 7,
        "2025_nr_permit": 8,
        "total permits": 9,
    }

    rows = []
    last_data_row = None

    def parse_labeled_permit(text: str):
        if text is None:
            return None, None
        s = str(text).replace("\xa0", " ").strip()
        if ":" not in s:
            return None, None
        left, right = s.split(":", 1)
        label = left.strip().lower().replace(" ", "")
        value = right.strip()
        if label == "res":
            return "res", value
        if label == "nonres":
            return "non_res", value
        if label == "total":
            return "total", value
        return None, None

    def cleanup_permit_value(v: str):
        if v is None:
            return ""
        s = str(v).replace("\xa0", " ").strip()
        if ":" in s:
            left, right = s.split(":", 1)
            l = left.strip().lower().replace(" ", "")
            if l in {"res", "nonres", "total"}:
                return right.strip()
        return s
    empty_streak = 0
    scanned = 0
    for vals_tuple in ws.iter_rows(min_row=hr + 1, min_col=1, max_col=40, values_only=True):
        vals = list(vals_tuple)
        scanned += 1
        if scanned > 5000:
            break
        txt_nonempty = [v for v in vals if v is not None and str(v).strip()]
        if not txt_nonempty:
            empty_streak += 1
            if empty_streak >= 100:
                break
            continue
        empty_streak = 0

        def getv(keys, fallback_key=None):
            for k in keys:
                if k in idx:
                    v = vals[idx[k]]
                    s = "" if v is None else str(v).strip()
                    if s != "":
                        return s
            if fallback_key and fallback_key in fallback_positions:
                p = fallback_positions[fallback_key]
                if p < len(vals):
                    v = vals[p]
                    return "" if v is None else str(v).strip()
            return ""

        hunt_name = getv(["hunt_name", "hunt name"], "hunt_name")
        hunt_code = getv(["hunt-number", "hunt number"], "hunt-number")
        species = getv(["species"], "species")

        labeled_key, labeled_val = parse_labeled_permit(
            getv(
                [
                    "2025_res_permit",
                    "2025-res-permit",
                    "2025_nr_permit",
                    "2025-nr-permit",
                    "total permits",
                    "total_permits",
                    "permits",
                ],
                "2025_res_permit",
            )
        )

        # Continuation row pattern: permit values on separate line.
        # Some workbooks carry merged hunt_code into the continuation row.
        is_continuation = False
        if labeled_key and last_data_row is not None:
            if not hunt_name and not species and (not hunt_code or hunt_code == last_data_row.get("hunt_code", "")):
                is_continuation = True
            elif (
                labeled_key in {"non_res", "total"}
                and hunt_code
                and hunt_code == last_data_row.get("hunt_code", "")
                and not hunt_name
            ):
                is_continuation = True

        if is_continuation:
            if labeled_key == "res":
                last_data_row["permits_2025_res"] = labeled_val
            elif labeled_key == "non_res":
                last_data_row["permits_2025_nr"] = labeled_val
            elif labeled_key == "total":
                last_data_row["permits_2025_total"] = labeled_val
            last_data_row["res_permits"] = last_data_row.get("permits_2025_res", "")
            last_data_row["non_res_permits"] = last_data_row.get("permits_2025_nr", "")
            last_data_row["total_permits"] = last_data_row.get("permits_2025_total", "")
            continue

        # Skip rows that clearly are not data rows
        if not hunt_name and not hunt_code and not species:
            continue

        row = {
            "hunt_name": hunt_name,
            "hunt_code": hunt_code,
            "sex_type": getv(["sex-type", "sex type", "sex"], "sex-type"),
            "species": species,
            "weapon": getv(["weapon"], "weapon"),
            "hunt_type": getv(["hunt-type", "hunt type"], "hunt-type"),
            "season": getv(["season", "season_dates", "season dates"], "season"),
            "permits_2025_res": getv(["2025_res_permit", "2025 resident permit", "2025 resident permits"], "2025_res_permit"),
            "permits_2025_nr": getv(["2025_nr_permit", "2025-nr-permit", "2025 nonresident permit", "2025 nonresident permits"], "2025_nr_permit"),
            "permits_2025_total": getv(["total permits", "total_permits", "permits total", "permits"], "total permits"),
            "source_file": path.name,
            "source_sheet": ws.title,
        }
        # Apply labeled permit override when encoded as "Res: X", "NonRes: Y", "Total: Z".
        if labeled_key == "res":
            row["permits_2025_res"] = labeled_val
        elif labeled_key == "non_res":
            row["permits_2025_nr"] = labeled_val
        elif labeled_key == "total":
            row["permits_2025_total"] = labeled_val
        # Normalize permit text fields in all cases.
        row["permits_2025_res"] = cleanup_permit_value(row["permits_2025_res"])
        row["permits_2025_nr"] = cleanup_permit_value(row["permits_2025_nr"])
        row["permits_2025_total"] = cleanup_permit_value(row["permits_2025_total"])
        # Explicit column names requested for downstream library/UI use.
        row["res_permits"] = row["permits_2025_res"]
        row["non_res_permits"] = row["permits_2025_nr"]
        row["total_permits"] = row["permits_2025_total"]
        rows.append(row)
        last_data_row = row

    return rows


def main():
    all_rows = []
    for fname in SOURCE_FILES:
        p = XLSX_DIR / fname
        if not p.exists():
            continue
        wb = load_workbook(p, read_only=True, data_only=True)
        for ws in wb.worksheets:
            all_rows.extend(parse_sheet(p, ws))

    # De-dup exact by hunt_code+hunt_name+source_file
    seen = set()
    deduped = []
    for r in all_rows:
        k = (r["hunt_code"], r["hunt_name"], r["source_file"])
        if k in seen:
            continue
        seen.add(k)
        deduped.append(r)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CANON)
        w.writeheader()
        w.writerows(deduped)

    print(f"NORMALIZED_ROWS={len(deduped)}")
    print(f"OUTPUT={OUT}")

    # Post-step: force permit cells to numeric types in raw 2026 xlsx files.
    # Missing permit data remains blank by rule.
    numeric_script = REPO / "pipeline" / "scripts" / "enforce_numeric_permit_cells.py"
    if numeric_script.exists():
        subprocess.run([sys.executable, str(numeric_script)], check=False)


if __name__ == "__main__":
    main()
