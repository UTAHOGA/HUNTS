#!/usr/bin/env python3
"""
Ensure permit columns are numeric cell types in 2026 raw XLSX files.

Rules:
- Keep missing values blank (None), never synthesize 0.
- Convert integer-like permit values to int cell types.
- Leave non-numeric text untouched.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


REPO = Path(r"D:\DOCUMENTS\GitHub\HUNTS")
XLSX_DIR = REPO / "pipeline" / "raw" / "2026" / "xlsx"

PERMIT_HEADERS = {
    "permits_res_2026",
    "permits_non-res_2026",
    "permits_total_2026",
    "permits_resident_2026",
    "permits_nonresident_2026",
    "permits_total",
}
CODE_HEADERS = {"hunt_code", "hunt_codes", "hunt_number"}


def convert_workbook(path: Path) -> int:
    wb = load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers_lc = [str(h).strip().lower() if h is not None else "" for h in headers]

    hunt_col = None
    permit_cols = []
    for i, h in enumerate(headers_lc, start=1):
        if h in CODE_HEADERS and hunt_col is None:
            hunt_col = i
        if h in PERMIT_HEADERS:
            permit_cols.append(i)

    if not permit_cols:
        return 0

    changes = 0
    for r in range(2, ws.max_row + 1):
        if hunt_col:
            code = ws.cell(r, hunt_col).value
            if code is None or str(code).strip() == "":
                continue

        for c in permit_cols:
            cell = ws.cell(r, c)
            v = cell.value
            if v is None:
                continue
            s = str(v).strip()
            if s == "":
                cell.value = None
                changes += 1
                continue
            try:
                n = int(float(s)) if s.endswith(".0") else int(s)
                cell.value = n
                changes += 1
            except Exception:
                # Preserve non-numeric text exactly as-is.
                pass

    if changes:
        wb.save(path)
    return changes


def main() -> int:
    changed = []
    for p in sorted(XLSX_DIR.glob("*.xlsx")):
        if p.name.startswith("~$"):
            continue
        try:
            c = convert_workbook(p)
            if c:
                changed.append((p.name, c))
        except Exception as exc:
            print(f"[WARN] {p.name}: {exc}")

    print(f"FILES_CHANGED={len(changed)}")
    for name, c in changed:
        print(f"{name}\tPERMIT_CELL_TYPE_FIXES={c}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
