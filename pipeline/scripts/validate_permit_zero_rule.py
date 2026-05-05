#!/usr/bin/env python3
"""
Hard rule: permit fields must stay blank when source data is missing.

This validator fails if it finds likely synthetic zero fills:
- permits_total* == 0 while permits_res* and permits_non* are both blank
- permits_res* == 0 while permits_non* and permits_total* are blank
- permits_non* == 0 while permits_res* and permits_total* are blank

Explicit source zeros are still allowed when paired with real data.
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


def _norm(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _find_col(headers_lc: List[str], candidates: List[str]) -> Optional[int]:
    for i, h in enumerate(headers_lc):
        if h in candidates:
            return i
    return None


def _is_zero(s: str) -> bool:
    return s in {"0", "0.0"}


def _validate_rows(
    rows: List[List[str]], headers: List[str], path: Path
) -> List[Tuple[int, str, str, str, str]]:
    headers_lc = [h.lower().strip() for h in headers]
    code_i = _find_col(headers_lc, ["hunt_code", "hunt_codes", "hunt_number"])
    res_i = _find_col(
        headers_lc, ["permits_res_2026", "permits_resident_2026", "permits_res"]
    )
    non_i = _find_col(
        headers_lc,
        ["permits_non-res_2026", "permits_nonresident_2026", "permits_non_res"],
    )
    tot_i = _find_col(headers_lc, ["permits_total_2026", "permits_total"])

    if code_i is None or (res_i is None and non_i is None and tot_i is None):
        return []

    findings: List[Tuple[int, str, str, str, str]] = []
    for idx, row in enumerate(rows, start=2):
        code = _norm(row[code_i]) if code_i < len(row) else ""
        if not code:
            continue
        res = _norm(row[res_i]) if res_i is not None and res_i < len(row) else ""
        non = _norm(row[non_i]) if non_i is not None and non_i < len(row) else ""
        tot = _norm(row[tot_i]) if tot_i is not None and tot_i < len(row) else ""

        if _is_zero(tot) and res == "" and non == "":
            findings.append((idx, code, res, non, tot))
        if _is_zero(res) and non == "" and tot == "":
            findings.append((idx, code, res, non, tot))
        if _is_zero(non) and res == "" and tot == "":
            findings.append((idx, code, res, non, tot))

    return findings


def _read_xlsx(path: Path) -> Tuple[List[str], List[List[str]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [_norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows: List[List[str]] = []
    for r in range(2, ws.max_row + 1):
        vals = [_norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if any(v != "" for v in vals):
            rows.append(vals)
    return headers, rows


def _read_csv(path: Path) -> Tuple[List[str], List[List[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        data = list(reader)
    if not data:
        return [], []
    headers = [_norm(x) for x in data[0]]
    rows = [[_norm(x) for x in r] for r in data[1:] if any(_norm(x) for x in r)]
    return headers, rows


def main() -> int:
    base = Path(r"D:\DOCUMENTS\GitHub\HUNTS\pipeline\raw\2026")
    xlsx_dir = base / "xlsx"
    csv_dir = base / "csv"

    files = sorted(
        [p for p in xlsx_dir.glob("*.xlsx") if not p.name.startswith("~$")]
        + [p for p in csv_dir.glob("*.csv")]
    )

    all_findings: Dict[Path, List[Tuple[int, str, str, str, str]]] = {}
    for path in files:
        try:
            if path.suffix.lower() == ".xlsx":
                headers, rows = _read_xlsx(path)
            else:
                headers, rows = _read_csv(path)
        except Exception as exc:
            print(f"[WARN] Could not read {path}: {exc}")
            continue
        findings = _validate_rows(rows, headers, path)
        if findings:
            all_findings[path] = findings

    if not all_findings:
        print("PASS: no synthetic permit zero fills found.")
        return 0

    print("FAIL: synthetic permit zero fills detected.")
    for path, findings in all_findings.items():
        print(f"\n{path}")
        for row_num, code, res, non, tot in findings[:50]:
            print(
                f"  row={row_num} hunt_code={code} permits_res={res!r} permits_nonres={non!r} permits_total={tot!r}"
            )
        if len(findings) > 50:
            print(f"  ... and {len(findings) - 50} more")
    return 1


if __name__ == "__main__":
    sys.exit(main())
